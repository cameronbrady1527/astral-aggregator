#!/usr/bin/env python3
# ==============================================================================
# export_changes_to_excel.py — Export change detection data to Excel format
# ==============================================================================
# Purpose: Read change files and export them to Excel workbook with separate sheets
# Sections: Imports, Public Exports, ChangesExporter Class, Main Function
# ==============================================================================

# ==============================================================================
# Imports
# ==============================================================================

# Standard Library -----
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Optional

# Third-Party -----
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Internal -----
# Add the app directory to the path
sys.path.insert(0, str(Path(__file__).parent.parent / "app"))
from utils.config import ConfigManager

# ==============================================================================
# Public exports
# ==============================================================================
__all__ = [
    'ChangesExporter',
    'main'
]

# ==============================================================================
# ChangesExporter Class
# ==============================================================================

class ChangesExporter:
    """Export change detection data to Excel format."""
    
    def __init__(self, changes_dir: str = "changes", output_dir: str = "output"):
        """Initialize the exporter with directories and styling."""
        # use project root paths instead of relative to scripts directory
        project_root = Path(__file__).parent.parent
        self.changes_dir = project_root / changes_dir
        self.output_dir = project_root / output_dir
        self.output_dir.mkdir(exist_ok=True)
        
        # load site configurations
        self.config_manager = ConfigManager()
        
        # excel styling
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def get_most_recent_changes_file(self, site_name: str) -> Optional[Path]:
        """Get the most recent changes file for a site by modification time."""
        if not self.changes_dir.exists():
            print(f"⚠️ Changes directory not found: {self.changes_dir}")
            return None
        
        # find all change files for this site
        pattern = f"{site_name}_*_changes.json"
        change_files = list(self.changes_dir.glob(pattern))
        
        if not change_files:
            print(f"⚠️ No change files found for site: {site_name}")
            return None
        
        # sort by modification time and get the most recent
        most_recent = max(change_files, key=lambda x: x.stat().st_mtime)
        print(f"📄 Found most recent changes file for {site_name}: {most_recent.name}")
        
        return most_recent
    
    def load_changes_data(self, file_path: Path) -> Dict[str, Any]:
        """Load changes data from JSON file with error handling."""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            return data
        except Exception as e:
            print(f"[ X ] Error loading changes file {file_path}: {e}")
            return {}
    
    def format_datetime(self, iso_datetime: str) -> str:
        """Format ISO datetime string to readable format with timezone handling."""
        try:
            dt = datetime.fromisoformat(iso_datetime.replace('Z', '+00:00'))
            return dt.strftime("%Y-%m-%d %H:%M:%S")
        except:
            return iso_datetime
    
    def create_dataframe(self, changes_data: Dict[str, Any]) -> pd.DataFrame:
        """Create a pandas DataFrame from changes data with field extraction."""
        if not changes_data or 'changes' not in changes_data:
            return pd.DataFrame()
        
        changes = changes_data['changes']
        if not changes:
            return pd.DataFrame()
        
        # extract relevant fields
        rows = []
        for change in changes:
            row = {
                'URL': change.get('url', ''),
                'Change Type': change.get('change_type', ''),
                'Detection Date & Time': self.format_datetime(change.get('detected_at', '')),
                'Details': change.get('details', ''),
                'Previous Hash': change.get('previous_hash', ''),
                'New Hash': change.get('new_hash', ''),
                'File Type': change.get('file_type', '')
            }
            rows.append(row)
        
        return pd.DataFrame(rows)
    
    def apply_excel_styling(self, worksheet, df: pd.DataFrame):
        """Apply styling to the Excel worksheet."""
        # apply borders to all cells
        for row in worksheet.iter_rows(min_row=1, max_row=len(df) + 1, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.border = self.border
        
        # style header row
        for cell in worksheet[1]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # set column width (max 50 characters)
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def export_site_changes(self, site_name: str, workbook: Workbook) -> bool:
        """Export changes for a specific site to a worksheet."""
        # get most recent changes file
        changes_file = self.get_most_recent_changes_file(site_name)
        if not changes_file:
            return False
        
        # load changes data
        changes_data = self.load_changes_data(changes_file)
        if not changes_data:
            return False
        
        # create DataFrame
        df = self.create_dataframe(changes_data)
        if df.empty:
            print(f"⚠️ No changes data found for {site_name}")
            return False
        
        # create worksheet
        sheet_name = site_name.replace(' ', '_')[:31]  # excel sheet name limit
        worksheet = workbook.create_sheet(title=sheet_name)
        
        # add data to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            worksheet.append(r)
        
        # apply styling
        self.apply_excel_styling(worksheet, df)
        
        # add metadata
        metadata = changes_data.get('metadata', {})
        detection_time = metadata.get('detection_time', 'Unknown')
        total_changes = metadata.get('total_changes', 0)
        
        # add summary info at the top
        worksheet.insert_rows(1, 3)
        worksheet['A1'] = f"Site: {site_name}"
        worksheet['A2'] = f"Detection Time: {self.format_datetime(detection_time)}"
        worksheet['A3'] = f"Total Changes: {total_changes}"
        
        # style summary rows
        for row_num in [1, 2, 3]:
            worksheet[f'A{row_num}'].font = Font(bold=True)
            worksheet[f'A{row_num}'].fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        print(f"✅ Exported {len(df)} changes for {site_name}")
        return True
    
    def export_all_changes(self) -> str:
        """Export changes for all sites to Excel."""
        print("🚀 Starting changes export to Excel...")
        
        # create workbook
        workbook = Workbook()
        
        # remove default sheet
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
        
        # get active sites
        active_sites = self.config_manager.get_active_sites()
        
        if not active_sites:
            print("⚠️ No active sites found in configuration")
            return ""
        
        exported_count = 0
        
        # export changes for each site
        for site_config in active_sites:
            site_name = site_config.name
            print(f"\n📊 Processing site: {site_name}")
            
            if self.export_site_changes(site_name, workbook):
                exported_count += 1
        
        if exported_count == 0:
            print("[ X ] No changes data found for any site")
            return ""
        
        # generate output filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = self.output_dir / f"changes_export_{timestamp}.xlsx"
        
        # save workbook
        workbook.save(output_file)
        
        print(f"\n✅ Successfully exported changes for {exported_count} sites")
        print(f"📄 Excel file saved: {output_file}")
        
        return str(output_file)


def main():
    """Run the main export function."""
    print("=" * 60)
    print("📊 Changes Export to Excel")
    print("=" * 60)
    
    # create exporter
    exporter = ChangesExporter()
    
    # export all changes
    output_file = exporter.export_all_changes()
    
    if output_file:
        print(f"\n🎉 Export completed successfully!")
        print(f"📁 Output file: {output_file}")
    else:
        print("\n[ X ] Export failed - no data found")
        sys.exit(1)


if __name__ == "__main__":
    main() 