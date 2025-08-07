#!/usr/bin/env python3
# ==============================================================================
# export_changes_cli.py — Command-line interface for exporting changes to Excel
# ==============================================================================
# Purpose: Provide user-friendly CLI for exporting change detection data
# Sections: Imports, Public Exports, Helper Functions, Main CLI
# ==============================================================================

# ==============================================================================
# Imports
# ==============================================================================

# Standard Library -----
import argparse
import sys
from datetime import datetime, timedelta
from pathlib import Path

# Internal -----
from export_changes_to_excel import ChangesExporter

# ==============================================================================
# Public exports
# ==============================================================================
__all__ = [
    'parse_date_range',
    'main'
]

# ==============================================================================
# Helper Functions
# ==============================================================================

def parse_date_range(date_range: str) -> tuple:
    """Parse date range string (e.g., '7d', '24h', '1w')."""
    if not date_range:
        return None, None
    
    try:
        value = int(date_range[:-1])
        unit = date_range[-1].lower()
        
        now = datetime.now()
        
        if unit == 'h':
            start_date = now - timedelta(hours=value)
        elif unit == 'd':
            start_date = now - timedelta(days=value)
        elif unit == 'w':
            start_date = now - timedelta(weeks=value)
        elif unit == 'm':
            start_date = now - timedelta(days=value * 30)
        else:
            raise ValueError(f"Invalid unit: {unit}")
        
        return start_date, now
    except (ValueError, IndexError):
        print(f"[ X ] Invalid date range format: {date_range}")
        print("   Use format like: 7d, 24h, 1w, 1m")
        sys.exit(1)

# ==============================================================================
# Main CLI
# ==============================================================================

def main():
    """Run the CLI interface for exporting changes to Excel."""
    parser = argparse.ArgumentParser(
        description="Export change detection data to Excel format",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Export all changes for all sites
  python export_changes_cli.py

  # Export changes for specific site
  python export_changes_cli.py --site "Judiciary UK"

  # Export changes from last 7 days
  python export_changes_cli.py --date-range 7d

  # Export changes for specific site from last 24 hours
  python export_changes_cli.py --site "Waverley Borough Council" --date-range 24h

  # Export with custom output filename
  python export_changes_cli.py --output "my_changes_report.xlsx"
        """
    )
    
    parser.add_argument(
        '--site', '-s',
        type=str,
        help='Export changes for specific site only'
    )
    
    parser.add_argument(
        '--date-range', '-d',
        type=str,
        help='Export changes from date range (e.g., 7d, 24h, 1w, 1m)'
    )
    
    parser.add_argument(
        '--output', '-o',
        type=str,
        help='Custom output filename'
    )
    
    parser.add_argument(
        '--list-sites',
        action='store_true',
        help='List all available sites and exit'
    )
    
    parser.add_argument(
        '--verbose', '-v',
        action='store_true',
        help='Verbose output'
    )
    
    args = parser.parse_args()
    
    # list sites if requested
    if args.list_sites:
        print("📋 Available sites:")
        exporter = ChangesExporter()
        active_sites = exporter.config_manager.get_active_sites()
        
        if not active_sites:
            print("   No active sites found in configuration")
        else:
            for site in active_sites:
                print(f"   • {site.name}")
        return
    
    print("=" * 60)
    print("📊 Changes Export to Excel (CLI)")
    print("=" * 60)
    
    # create exporter
    exporter = ChangesExporter()
    
    # parse date range
    start_date, end_date = parse_date_range(args.date_range) if args.date_range else (None, None)
    
    if start_date and end_date:
        print(f"📅 Date range: {start_date.strftime('%Y-%m-%d %H:%M')} to {end_date.strftime('%Y-%m-%d %H:%M')}")
    
    # get active sites
    active_sites = exporter.config_manager.get_active_sites()
    
    if not active_sites:
        print("[ X ] No active sites found in configuration")
        sys.exit(1)
    
    # filter sites if specified
    if args.site:
        filtered_sites = [site for site in active_sites if site.name == args.site]
        if not filtered_sites:
            print(f"[ X ] Site '{args.site}' not found in active sites")
            print("   Available sites:")
            for site in active_sites:
                print(f"     • {site.name}")
            sys.exit(1)
        active_sites = filtered_sites
        print(f"🎯 Exporting changes for site: {args.site}")
    
    # create workbook
    from openpyxl import Workbook
    workbook = Workbook()
    
    # remove default sheet
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])
    
    exported_count = 0
    
    # export changes for each site
    for site_config in active_sites:
        site_name = site_config.name
        print(f"\n📊 Processing site: {site_name}")
        
        if exporter.export_site_changes(site_name, workbook):
            exported_count += 1
    
    if exported_count == 0:
        print("[ X ] No changes data found for any site")
        sys.exit(1)
    
    # generate output filename
    project_root = Path(__file__).parent.parent
    if args.output:
        output_file = project_root / "output" / args.output
        if not output_file.suffix:
            output_file = output_file.with_suffix('.xlsx')
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = project_root / "output" / f"changes_export_{timestamp}.xlsx"
    
    # ensure output directory exists
    output_file.parent.mkdir(exist_ok=True)
    
    # save workbook
    workbook.save(output_file)
    
    print(f"\n✅ Successfully exported changes for {exported_count} sites")
    print(f"📄 Excel file saved: {output_file}")
    
    if args.verbose:
        print(f"\n📋 File details:")
        print(f"   Size: {output_file.stat().st_size:,} bytes")
        print(f"   Created: {datetime.fromtimestamp(output_file.stat().st_mtime).strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"   Sheets: {', '.join(workbook.sheetnames)}")


if __name__ == "__main__":
    main() 